# Revision History
#
#
# Authors: Elvis Segbeaya
# Dealing with WellData in general. Getting File ready for different Use Case Scenarios
#

import csv
import json
import logging
import time
from datetime import datetime as dt, date, datetime, timedelta
import pandas as pd
import SampleHelper
import welldataAPI
from retry import retry
from tenacity import retry, stop_after_attempt, wait_fixed
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl import load_workbook
from openpyxl.formatting.rule import IconSet, FormatObject, Rule, CellIsRule, IconSetRule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Font, Border
import re
import numpy as np
from pydantic import BaseModel
from pydantic.dataclasses import dataclass
import schedule
import time
KeepCacheForDevelopment = False


# creating a logging file

# come here to edit the attributes that you want to hold.
# get jobID, rigNumber, WellName, Startdate, EndDate #adding owner since we should verify who the owner is

class Well(BaseModel):
    jobID = ""
    jobName = ""
    Owner = ""
    rigNumber = ""
    Startdate = ""
    EndDate = ""



class UnitV1(BaseModel):
    id: str
    name: str
    abbreviation: str


class Attribute(BaseModel):
    id: str
    mode: str


def thresholdCheck(min=0, max=150000, input=1):
    if input > min and input <= max:
        print(f'{input} is within range')
        return True
    else:
        print(f'{input} is not within range of {min} - {max}')
        return False

# Set up logging
logging.basicConfig(filename='RAEAutomation.log', level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s')

def main():
    logging.info('Starting script')
    #######################################################################
    # Setup
    #######################################################################

    # SampleHelper file is used to setup the configuration
    SampleHelper.SetupLogging(logging, logging.DEBUG)
    SampleHelper.SetupLocale()

    SampleHelper.VersionCheck()

    # WellData Config
    configFile = SampleHelper.SetConfigFile(welldataAPI.defaultConfig())

    CFGdefault = SampleHelper.readConfig(configFile, welldataAPI.defaultConfig())

    # parses the api request -> this currently uses the welldata.cfg as the config file
    CFG = SampleHelper.readConfig(configFile, welldataAPI.serverConfig('welldata net'))
    CFG.update(CFGdefault)

    # Process Name, used for stateless process control
    today = date.today().strftime('%m-%d-%Y')
    today2 = date.today().strftime('%Y-%m-%d')
    datetime_string_from = today2 + 'T06:05:17'
    datetime_string_to = today2 + 'T06:06:17'
    print(today)
    if CFG['FromHours'] != '':
        datetime_string_from = CFG['FromHours']

    if CFG['ToHours'] != '':
        datetime_string_to = CFG['ToHours']


    print(datetime_string_from)
    print(datetime_string_to)
    processName = 'RAE Reports'

    DataSource = None
    if CFG['APIUrl'] == 'https://data.welldata.net/api/v1':
        DataSource = 'WellData Data API .NET'
    else:
        logging.error("Unknown Datasource {}. Aborting".format(DataSource))
        quit()

    URLs_v1 = welldataAPI.URLs_v1(CFG['APIUrl'], CFG['ContractorName'], CFG['OperatorName'], CFG['JobStatus'])

    # Generates Token for Header
    # switch jobs based on ca or US
    # token2 = welldataAPI.getToken(CFG['APIUrl'], CFG['appID'], CFG['rae_ca_user'], CFG['rae_ca_password'])

    #US jobs
    token = welldataAPI.getToken(CFG['APIUrl'], CFG['appID_rae'], CFG['rae_user'], CFG['rae_password'])

    # #Canada Jobs
    # token = welldataAPI.getToken(CFG['APIUrl'], CFG['appID_rae_ca'], CFG['rae_ca_user'], CFG['rae_ca_password'])


    # ONLY Get Patterson wells - getWells will filter them for us



    # # Writing DataFrame to Excel sheet
    # writer = pd.ExcelWriter(f'Rig List.xlsx', engine='openpyxl')
    # df = pd.DataFrame(jobs)
    # df.to_excel(writer, sheet_name='Jobs Processed', index=False)
    # writer.close()
    ######################################################################
    # Main Code- RAE Report Stuff below, API Configuration Stuff above
    ######################################################################


    # Variables

    # Define emojis as Unicode characters
    emoji_check = 3  # u'\u2705'  # ✅
    emoji_exclamation = 2  # u'\u2757'  # ❗
    emoji_x = 0  # u'\u274C'  # ❌

    attributeList = ['HookLoad', 'PumpPressure', 'BlockHeight', 'PumpSpm', 'PumpSpm2', 'RotaryTorque', 'BitPosition', 'SlipStatus']
    operators = ['Coterra', 'Upcurve', 'Alchemist', 'Endeavor', 'DB4', 'Piendeda', 'Surge', 'Black Swan', 'ConocoPhillips']
    processedJobList = []
    RAE_Operators = CFG['RAE_OperatorName']
    RAE_Rigs = CFG['RAE_Rigs']
    RAEJobs = []
    MR_Report_Ids = []
    MR_Report_Comments = []
    expectedAttr = CFG['ChannelsToOutput']
    attributeList = CFG['ChannelsToOutput']
    tmpJobs = []

    jobsTimeBased = []
    count = 1
    jobcount = 1
    RAEchoice = CFG['ActiveRAEJobsOnly']

    report = ''
    well = ''
    wellList = []
    ZeroReportList = []


    # Creating a lookup table of rigs
    lookup_table = {}
    tmpAllJobs = welldataAPI.getJobs(URLs_v1['getJobs'], token, CFG, take=1000, total=False, jobStatus="ActiveJobs")
    for w in tmpAllJobs:
        #jobs.append([w['id'], w['assetInfoList'][0]['owner'], w['assetInfoList'][0]['name']])
        key = f"{w['assetInfoList'][0]['owner']} {w['assetInfoList'][0]['name']}"
        # key = w['siteInfoList'][0]['owner']
        value = w['id']
        lookup_table[key] = value

    # Creating our Skinny table
    # Print the lookup table
    print("Lookup Table:")
    for key, value in lookup_table.items():
        print(f"Key: {key} | Value: {value}")


    for w in RAE_Rigs:
        for key, value in lookup_table.items():
            if w in key:
                tmpJobs.append(value)
    print("Temp Jobs:")
    for t in tmpJobs:
        print(t)

    try:
        #for w in tmpJobs:#
        for w in tmpAllJobs:
            #well = str(w)
            well = w['id']
            attsLst = []


            # variables:
            holder = []
            HookLoadbool = emoji_x
            PumpPressurebool = emoji_x
            BlockHeightbool = emoji_x
            PumpSpmbool = emoji_x
            PumpSpm2bool = emoji_x
            PumpSpm3bool = emoji_x
            RotaryTorquebool = emoji_x
            BitPositionbool = emoji_x
            BitStatusbool = emoji_x
            SlipStatusbool = emoji_x
            tpDriveRPM = emoji_x
            comment = 'NA'
            comment24 = 'NA'
            reportDate = ''
            reportID = ''
            reportStatus = ''
            realTime = emoji_x
            tpDriveTorq = emoji_x
            weightonBit = emoji_x
            RP_Fast = emoji_x
            tHookLoad = emoji_x




            # Attribute Values
            HookLoad_val = ''
            PumpPressure_val = ''
            BlockHeight_val = ''
            PumpSpm_val = ''
            PumpSpm2_val = ''
            PumpSpm3_val = ''
            RotaryTorque_val = ''
            tpDriveRPM_val = ''
            tpDriveTorq_val = ''
            weightonBit_val = ''
            BitPosition_val = ''
            BitStatus_val = ''
            RP_Fast_val = ''
            SlipStatus_val = ''
            tHookLoad_val = ''
            Iadc_Rig_val = ''
            Iadc2_Rig_val = ''
            Iadc3_Rig_val = ''



            # Checking for real time data capability
            rTime = welldataAPI.getApiCall(URLs_v1['getJobsIdCapabilities'], token, CFG, jobId=well)
            if 'realTime' in str(rTime):
                if rTime[0]['realTime'] == 'Supported':
                    realTime = emoji_check


            # 2 Get Attribute for job
            q = welldataAPI.getApiCall(URLs_v1['getAttributes'], token, CFG, jobId=well)

            # for c in q[0]['attributes']:
            #     print(c['id'])

            # append attribute fields to holder
            for c in q[0]['attributes']:
                if len(c) == 0:
                    check = ''
                # print(f'processing job {count} of {len(q[0]["attributes"])} for jobid: {w["id"]} with ID: {c['alias']['witsml_mnemonic']} and has data equal to : {c["hasData"]}')
                count = count + 1
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'HOOKLOAD_MAX': # hookload
                    #HookLoadbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'STP_PRS_1': #'PumpPressure':
                    #PumpPressurebool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'BLOCK_POS': #Blockheight
                    #BlockHeightbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'MP1_SPM': #'PumpSpm'
                    #PumpSpmbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'MP2_SPM': #'PumpSpm2'
                    #PumpSpm2bool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'MP3_SPM': #'PumpSpm3'
                    #PumpSpm3bool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'ROT_TORQUE': # 'RotaryTorque':
                    #RotaryTorquebool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'TD_SPEED':  # tpDriveRPM
                    #tpDriveRPM = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'TD_TORQUE':  # tpDriveTorq
                    #tpDriveTorq = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'WOB':  # WOB
                    #weightonBit = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'BIT_DEPTH':  # BitPosition
                    #BitPositionbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'BIT_ON_BTM':  # BitStatus
                    #BitStatusbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'FAST_ROP_FT_HR': #'FastRopFtHr':  # ROP-F
                    #RP_Fast = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'SLIPS_STAT': #'SlipStatus':  # SlipStatus
                    #SlipStatusbool = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'Trigger Hkld' : #'TrigHkld'
                    #tHookLoad = emoji_check
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'IADC_RIG_ACTIVITY': #'FastRopFtHr':  # ROP-F
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'IADC_RIG_ACTIVITY2': #'SlipStatus':  # SlipStatus
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)
                if c['hasData'] == True and c['alias']['witsml_mnemonic'] == 'IADC_RIG_ACTIVITY3' : #'TrigHkld'
                    attr = Attribute(id=c['id'], mode='Last')
                    attsLst.append(attr)

            if len(attsLst)== 0:
                # Appending jobs to dataFrame
                job = welldataAPI.getJobs(URLs_v1['getJobsId'], token, CFG, take=1000, total=False, jobStatus="ActiveJobs", jobId=well)
                # Appending for RAEJobs
                holder.append(job[0]["assetInfoList"][0]["owner"])
                holder.append(job[0]["assetInfoList"][0]["name"])
                holder.append(job[0]['siteInfoList'][0]['owner'])
                holder.append(job[0]['name'])
                # holder.append(well)

                holder.append(str(Iadc_Rig_val))
                holder.append(str(Iadc2_Rig_val))
                holder.append(str(Iadc3_Rig_val))
                holder.append(int(HookLoadbool))
                holder.append(int(PumpPressurebool))
                holder.append(int(BlockHeightbool))
                holder.append(int(PumpSpmbool))
                holder.append(int(PumpSpm2bool))
                holder.append(int(PumpSpm3bool))
                holder.append(int(RotaryTorquebool))
                holder.append(int(tpDriveRPM))
                holder.append(int(tpDriveTorq))
                holder.append(int(weightonBit))
                holder.append(int(RP_Fast))
                holder.append(int(tHookLoad))
                holder.append(int(BitPositionbool))
                holder.append(str(BitStatusbool))
                holder.append(str(SlipStatusbool))

                holder.append(comment)
                holder.append(comment24)
                holder.append(reportID)
                holder.append(reportDate)
                # holder.append(reportStatus)
                count = count + 1
                jobcount = jobcount + 1

                RAEJobs.append(holder)
                continue
            else:
                to_time = datetime_string_to
                from_time = datetime_string_from
                formatted_to_time = datetime.fromisoformat(to_time)
                formatted_from_time = datetime.fromisoformat(from_time)
                hist_interval = CFG['HistoricInterval']
                hist_payload = welldataAPI.HistoricalTimeRequest(attributes=attsLst, toTime=to_time, fromTime=from_time, interval=hist_interval)
                hist = welldataAPI.historical_data_time(well, hist_payload.json(exclude_unset=True), token=token)
                jobsTimeBased.append([well, hist])

            attribute_mapping = {}

            if len(hist['timeRecords']) ==  0:
                # Appending jobs to dataFrame
                job = welldataAPI.getJobs(URLs_v1['getJobsId'], token, CFG, take=1000, total=False, jobStatus="ActiveJobs", jobId=well)
                # Appending for RAEJobs
                holder.append(job[0]["assetInfoList"][0]["owner"])
                holder.append(job[0]["assetInfoList"][0]["name"])
                holder.append(job[0]['siteInfoList'][0]['owner'])
                holder.append(job[0]['name'])
                # holder.append(well)

                holder.append(str(Iadc_Rig_val))
                holder.append(str(Iadc2_Rig_val))
                holder.append(str(Iadc3_Rig_val))
                holder.append(int(HookLoadbool))
                holder.append(int(PumpPressurebool))
                holder.append(int(BlockHeightbool))
                holder.append(int(PumpSpmbool))
                holder.append(int(PumpSpm2bool))
                holder.append(int(PumpSpm3bool))
                holder.append(int(RotaryTorquebool))
                holder.append(int(tpDriveRPM))
                holder.append(int(tpDriveTorq))
                holder.append(int(weightonBit))
                holder.append(int(RP_Fast))
                holder.append(int(tHookLoad))
                holder.append(int(BitPositionbool))
                holder.append(str(BitStatusbool))
                holder.append(str(SlipStatusbool))

                holder.append(comment)
                holder.append(comment24)
                holder.append(reportID)
                holder.append(reportDate)
                # holder.append(reportStatus)
                count = count + 1
                jobcount = jobcount + 1

                RAEJobs.append(holder)
                continue
            values_timestamp_0 = hist['timeRecords'][0]['values']

            # Iterate over the 'attributes' list and map each attribute to its value from timestamp 0
            for i, attribute in enumerate(hist['attributes']):
                attribute_id = attribute['id']
                attribute_value = values_timestamp_0[i][1]
                attribute_mapping[attribute_id] = attribute_value

            print(attribute_mapping)

            # Iterate over the items in the dictionary
            for key, value in attribute_mapping.items():
                if key == 'HookLoad':
                    HookLoad_val = value
                    if isinstance(HookLoad_val, float) or isinstance(HookLoad_val, int):
                        if HookLoad_val == 0.00:
                            HookLoadbool = emoji_exclamation
                        elif HookLoad_val > 0.00:
                            HookLoadbool = emoji_check
                        else:
                            HookLoadbool = emoji_x

                    # if isinstance(HookLoad_val, float):
                    #     if thresholdCheck(CFG['HookLoadbool_min'], CFG['HookLoadbool_max'], float(HookLoad_val)) == False:
                    #         HookLoadbool = emoji_exclamation
                    #     else:
                    #         HookLoadbool = emoji_check

                if key == 'PumpPressure':
                    PumpPressure_val = value
                    if isinstance(PumpPressure_val, float) or isinstance(PumpPressure_val, int) :
                        if PumpPressure_val == 0.00:
                            PumpPressurebool = emoji_exclamation
                        elif PumpPressure_val > 0.00:
                            PumpPressurebool = emoji_check
                        else:
                            PumpPressurebool = emoji_x

                    # if isinstance(PumpPressure_val, float):
                    #     if thresholdCheck(CFG['PumpPressurebool_min'], CFG['PumpPressurebool_max'], float(PumpPressure_val)) == False:
                    #         PumpPressurebool = emoji_exclamation
                    #     else:
                    #         PumpPressurebool = emoji_check

                if key == 'BlockHeight':
                    BlockHeight_val = value
                    if isinstance(BlockHeight_val, float) or isinstance(BlockHeight_val, int):
                        if BlockHeight_val == 0.00:
                            BlockHeightbool = emoji_exclamation
                        elif BlockHeight_val > 0.00:
                            BlockHeightbool = emoji_check
                        else:
                            BlockHeightbool = emoji_x

                    # if isinstance(BlockHeight_val, float):
                    #     if thresholdCheck(CFG['BlockHeightbool_min'], CFG['BlockHeightbool_max'], float(BlockHeight_val)) == False:
                    #         BlockHeightbool = emoji_exclamation
                    #     else:
                    #         BlockHeightbool = emoji_check

                if key == 'PumpSpm':
                    PumpSpm_val = value
                    if isinstance(PumpSpm_val, float) or isinstance(PumpSpm_val, int):
                        if PumpSpm_val == 0:
                            PumpSpmbool = emoji_exclamation
                        elif PumpSpm_val > 0:
                            PumpSpmbool = emoji_check
                        else:
                            PumpSpmbool = emoji_x
                    # if isinstance(PumpSpm_val, float):
                    #     if thresholdCheck(CFG['PumpSpmbool_min'], CFG['PumpSpmbool_max'], float(PumpSpm_val)) == False:
                    #         PumpSpmbool = emoji_exclamation
                    #     else:
                    #         PumpSpmbool = emoji_check

                if key == 'PumpSpm2':
                    PumpSpm2_val = value
                    if isinstance(PumpSpm2_val, float) or isinstance(PumpSpm2_val, int):
                        if PumpSpm2_val == 0:
                            PumpSpm2bool = emoji_exclamation
                        elif PumpSpm2_val > 0:
                            PumpSpm2bool = emoji_check
                        else:
                            PumpSpm2bool = emoji_x

                    # if isinstance(PumpSpm2_val, float):
                    #     if thresholdCheck(CFG['PumpSpm2bool_min'], CFG['PumpSpm2bool_max'], float(PumpSpm2_val)) == False:
                    #         PumpSpm2bool = emoji_exclamation
                    #     else:
                    #         PumpSpm2bool = emoji_check

                if key == 'PumpSpm3':
                    PumpSpm3_val = value
                    if isinstance(PumpSpm3_val, float) or isinstance(PumpSpm3_val, int):
                        if PumpSpm3_val == 0:
                            PumpSpm3bool = emoji_exclamation
                        elif PumpSpm3_val > 0:
                            PumpSpm3bool = emoji_check
                        else:
                            PumpSpm3bool = emoji_x
                    #
                    # if isinstance(PumpSpm3_val, float):
                    #     if thresholdCheck(CFG['PumpSpm3bool_min'], CFG['PumpSpm3bool_max'], float(PumpSpm3_val)) == False:
                    #         PumpSpm3bool = emoji_exclamation
                    #     else:
                    #         PumpSpm3bool = emoji_check

                if key == 'RotaryTorque':
                    RotaryTorque_val = value
                    if isinstance(RotaryTorque_val, int):
                        if RotaryTorque_val > 0:
                            RotaryTorquebool = emoji_check
                        elif RotaryTorque_val == 0:
                            RotaryTorquebool = emoji_exclamation
                        else:
                            RotaryTorquebool = emoji_x

                    # if isinstance(RotaryTorque_val, float):
                    #     if thresholdCheck(CFG['RotaryTorquebool_min'], CFG['RotaryTorquebool_max'], float(RotaryTorque_val)) == False:
                    #         RotaryTorquebool = emoji_exclamation
                    #     else:
                    #         RotaryTorquebool = emoji_check

                if key == 'TopDrvRpm':  # tpDriveRPM
                    tpDriveRPM_val = value
                    if isinstance(tpDriveRPM_val, float) or isinstance(tpDriveRPM_val, int):
                        if tpDriveRPM_val == 0:
                            tpDriveRPM = emoji_exclamation
                        elif tpDriveRPM_val > 0:
                            tpDriveRPM = emoji_check
                        else:
                            tpDriveRPM = emoji_x
                    # if isinstance(tpDriveRPM_val, float):
                    #     if thresholdCheck(CFG['tpDriveRPM_min'], CFG['tpDriveRPM_max'], float(tpDriveRPM_val)) == False:
                    #         tpDriveRPM = emoji_exclamation
                    #     else:
                    #         tpDriveRPM = emoji_check

                if key == 'TopDrvTorque':  # tpDriveTorq
                    tpDriveTorq_val = value
                    if isinstance(tpDriveTorq_val, float) or isinstance(tpDriveTorq_val, int):
                        if tpDriveTorq_val == 0:
                            tpDriveTorq = emoji_exclamation
                        elif tpDriveTorq_val > 0:
                            tpDriveTorq = emoji_check
                        else:
                            tpDriveTorq = emoji_x

                    # if isinstance(tpDriveTorq_val, float):
                    #     if thresholdCheck(CFG['tpDriveTorq_min'], CFG['tpDriveTorq_max'], float(tpDriveTorq_val)) == False:
                    #         tpDriveTorq = emoji_exclamation
                    #     else:
                    #         tpDriveTorq = emoji_check

                if key == 'BitWeightQualified':  # WOB
                    weightonBit_val = value
                    if isinstance(weightonBit_val, float) or isinstance(weightonBit_val, int):
                        if weightonBit_val == 0:
                            weightonBit = emoji_exclamation
                        elif weightonBit_val > 0:
                            weightonBit = emoji_check
                        else:
                            weightonBit = emoji_x

                    # if isinstance(tpDriveTorq_val, float):
                    #     if thresholdCheck(CFG['WOB_min'], CFG['WOB_max'], float(weightonBit_val)) == False:
                    #         weightonBit = emoji_exclamation
                    #     else:
                    #         weightonBit = emoji_check

                if key == 'BitPosition':  # BitPosition
                    BitPosition_val = value
                    if isinstance(BitPosition_val, float) or isinstance(BitPosition_val, int):
                        if BitPosition_val == 0:
                            BitPositionbool = emoji_exclamation
                        elif BitPosition_val > 0:
                            BitPositionbool = emoji_check
                        else:
                            BitPositionbool = emoji_x

                    # if isinstance(BitPosition_val, float):
                    #     if thresholdCheck(CFG['BitPositionbool_min'], CFG['BitPositionbool_max'], float(BitPosition_val)) == False:
                    #         BitPositionbool = emoji_exclamation
                    #     else:
                    #         BitPositionbool = emoji_check

                if key == 'BitStatus':  # BitStatus
                    BitStatus_val = value
                    if BitStatus_val == 0:
                        BitStatusbool = 'On'
                    elif BitStatus_val == 1:
                        BitStatusbool = 'Off'
                    else:
                        BitStatusbool = emoji_x
                    # if isinstance(BitStatus_val, float):
                    #     if thresholdCheck(CFG['BitStatusbool_min'], CFG['BitStatusbool_max'], float(BitStatus_val)) == False:
                    #         BitStatusbool = emoji_exclamation
                    #     else:
                    #         BitStatusbool = emoji_check

                if key == 'FastRopFtHr':  # ROP-F
                    RP_Fast_val = value
                    if isinstance(RP_Fast_val, float) or isinstance(RP_Fast_val, int):
                        if RP_Fast_val == 0:
                            RP_Fast = emoji_exclamation
                        elif RP_Fast_val > 0:
                            RP_Fast = emoji_check
                        else:
                            RP_Fast = emoji_x

                    # if isinstance(RP_Fast_val, float):
                    #     if thresholdCheck(CFG['RP_Fast_min'], CFG['RP_Fast_max'], float(RP_Fast_val)) == False:
                    #         RP_Fast = emoji_exclamation
                    #     else:
                    #         RP_Fast = emoji_check

                if key == 'SlipStatus':  # SlipStatus
                    SlipStatus_val = value
                    if SlipStatus_val == 0:  # on
                        SlipStatusbool = 'Resetting'
                    elif SlipStatus_val == 1:  # on
                        SlipStatusbool = 'in'
                    elif SlipStatus_val == 2:  # off
                        SlipStatusbool = 'out'
                    else:
                        SlipStatusbool = emoji_x
                    # if isinstance(SlipStatus_val, float):
                    #     if thresholdCheck(CFG['SlipStatusbool_min'], CFG['SlipStatusbool_max'], float(SlipStatus_val)) == False:
                    #         SlipStatusbool = emoji_exclamation
                    #     else:
                    #         SlipStatusbool = emoji_check

                if key == 'TrigHkld':  # T-HL
                    tHookLoad_val = value
                    if isinstance(tHookLoad_val, float) or isinstance(tHookLoad_val, int):
                        if tHookLoad_val > 0:
                            tHookLoad = emoji_check
                        elif tHookLoad_val == 0:
                            tHookLoad = emoji_exclamation
                        else:
                            tHookLoad = emoji_x

                        # if thresholdCheck(CFG['tHookLoad_min'], CFG['tHookLoad_max'], float(tHookLoad_val)) == False:
                        #     tHookLoad = emoji_exclamation
                        # else:
                        #     tHookLoad = emoji_check

                if key == 'IadcRigActivity':  # IADC_RIG_ACTIVITY
                    Iadc_Rig_val = value
                if key == 'IadcRigActivity2':  # IADC_RIG_ACTIVITY
                    Iadc2_Rig_val = value
                if key == 'RigActivity':  # IADC_RIG_ACTIVITY
                    Iadc3_Rig_val = value



            # getting the comment portion


            r = welldataAPI.getReports(URLs_v1['getReportsClassificationReportGroup'], token, CFG, jobId=well, reportGroupId=2, classification='daily')
            MR_Report_Ids.append([well, r])
            # Checking iif the morning reports exists for job ID
            if well in ZeroReportList:
                continue
            elif len(r) == 0:
                ZeroReportList.append(well)
                comment = 'no morning report'
                # for z in ZeroReportList:
                #     print(z)
                # continue


            else:
                # Go through Id's and pull pdf downloads of reports
                report = welldataAPI.getReports(URLs_v1['getReportsClassificationReportGroupFileFormat'], token, CFG, jobId=well, reportGroupId=2, classification='daily',fileFormat='JSON', reportId=r[0]['id'])
                if len(report) == 0:
                    ZeroReportList.append(well)
                    comment = 'no morning report'

                if 'GenericAmericanMorningReportDW' in str(report):
                    # continue
                    reportDate = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['Header']['Date']
                    comment = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['Header']['OpsAtReportTime']
                    if 'OpsNext24' in str(report):
                        comment24 = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['Header']['OpsNext24']
                    reportID = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['GenericAmericanMorningReportDW']['ReportAttributes']['ReportStatus']
                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['GenericAmericanMorningReportDW']['ActivityDetails']['Items']:
                        if 'ActCode' in rep:
                            MR_Report_Comments.append([well, r[0]['id'], r[0]['date'], rep['ActCode'], rep['DescriptionOfWork']])
                        else:
                            continue
                    # 'HandPMorningReport'
                elif 'HandPMorningReport' in str(report):
                    reportDate = report[0]['Reports'][0]['HandPMorningReport']['Header']['Date']
                    comment = report[0]['Reports'][0]['HandPMorningReport']['Operations']['PresentOp']
                    comment24 = ''
                    reportID = report[0]['Reports'][0]['HandPMorningReport']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['HandPMorningReport']['ReportAttributes']['ReportStatus']
                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['HandPMorningReport']['TimeSummary']['Items']:
                        if 'ActivityCode' in rep:
                            MR_Report_Comments.append(
                                [well, r[0]['id'], r[0]['date'], rep['ActivityCode'], rep['ActivityDetails']])
                        else:
                            continue
                    # 'ScanMorningReport'
                elif 'ScanMorningReport' in str(report):
                    reportDate = report[0]['Reports'][0]['ScanMorningReport']['Header']['Date']
                    reportID = report[0]['Reports'][0]['ScanMorningReport']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['ScanMorningReport']['ReportAttributes']['ReportStatus']
                    comment = report[0]['Reports'][0]['ScanMorningReport']['Header']['PresentOp']
                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['ScanMorningReport']['TimeBreakDown']['Items']:
                        if 'ActivityCode' in rep:
                            MR_Report_Comments.append(
                                [well, r[0]['id'], r[0]['date'], rep['ActivityCode'], rep['ActivityDetails']])
                        else:
                            continue

                    # 'RapadMorningReport'
                elif 'RapadMorningReport' in str(report):
                    reportDate = report[0]['Reports'][0]['RapadMorningReport']['Header']['ReportDate']
                    comment = report[0]['Reports'][0]['RapadMorningReport']['Header']['OperationsActivityCurrent']
                    comment24 = report[0]['Reports'][0]['RapadMorningReport']['Header']['OperationsActivityNext24Hours']
                    reportID = report[0]['Reports'][0]['RapadMorningReport']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['RapadMorningReport']['ReportAttributes']['ReportStatus']

                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['RapadMorningReport']['ActivityDetails']['Items']:
                        if 'OperationsActivityCode' in rep:
                            MR_Report_Comments.append([well, r[0]['id'], r[0]['date'], rep['OperationsActivityCode'], rep['OperationsActivityDescription']])
                        else:
                            continue
                    # 'PattersonMorningReportRevB'
                elif 'PattersonMorningReportRevB' in str(report):
                    reportDate = report[0]['Reports'][0]['PattersonMorningReportRevB']['Header']['ReportDate']
                    comment = report[0]['Reports'][0]['PattersonMorningReportRevB']['OperationsCasingDetails']['operations_at_report_time']
                    if 'operations_next_24_hours' in report[0]['Reports'][0]['PattersonMorningReportRevB']['OperationsCasingDetails']:
                        comment24 = report[0]['Reports'][0]['PattersonMorningReportRevB']['OperationsCasingDetails']['operations_next_24_hours']
                    reportID = report[0]['Reports'][0]['PattersonMorningReportRevB']['ReportAttributes']['ReportID']
                    reportStatus = report[0]['Reports'][0]['PattersonMorningReportRevB']['ReportAttributes']['ReportStatus']

                    # Getting the Activity code and comments
                    for rep in report[0]['Reports'][0]['PattersonMorningReportRevB']['ActivityDetails']['Items']:
                        if 'details' in rep:
                            if 'code' in rep:
                                code = rep['code']
                            else:
                                code = ''
                            MR_Report_Comments.append(
                                [well, r[0]['id'], r[0]['date'], code, rep['details']])
                        else:
                            continue
                else:
                    processedJobList.append([well, str(report[0]['Reports'][0])])
                    # print(processedJobList)
                    continue


            # Appending jobs to dataFrame
            job = welldataAPI.getJobs(URLs_v1['getJobsId'], token, CFG, take=1000, total=False, jobStatus="ActiveJobs", jobId=well)
            # Appending for RAEJobs
            holder.append(job[0]["assetInfoList"][0]["owner"])
            holder.append(job[0]["assetInfoList"][0]["name"])
            holder.append(job[0]['siteInfoList'][0]['owner'])
            holder.append(job[0]['name'])
            #holder.append(well)

            holder.append(str(Iadc_Rig_val))
            holder.append(str(Iadc2_Rig_val))
            holder.append(str(Iadc3_Rig_val))
            holder.append(int(HookLoadbool))
            holder.append(int(PumpPressurebool))
            holder.append(int(BlockHeightbool))
            holder.append(int(PumpSpmbool))
            holder.append(int(PumpSpm2bool))
            holder.append(int(PumpSpm3bool))
            holder.append(int(RotaryTorquebool))
            holder.append(int(tpDriveRPM))
            holder.append(int(tpDriveTorq))
            holder.append(int(weightonBit))
            holder.append(int(RP_Fast))
            holder.append(int(tHookLoad))
            holder.append(int(BitPositionbool))
            holder.append(str(BitStatusbool))
            holder.append(str(SlipStatusbool))

            holder.append(comment)
            holder.append(comment24)
            holder.append(reportID)
            holder.append(reportDate)
            # holder.append(reportStatus)
            count = count + 1
            jobcount = jobcount + 1

            RAEJobs.append(holder)



    except Exception as ex:
        wellList.append(well)
        print(well)
        print(report)
        logging.error("Error sending request to server")
        logging.error(f"Exception: {ex}")
        pass

    # Writing DataFrame to Excel sheet
    writer = pd.ExcelWriter(f'RAEJobsListTest {today} .xlsx', engine='openpyxl')
    header = ['Contractor', 'Rig Number', 'Operator', 'Well name', 'IADC','IADC 2','IADC3', 'HookLoad', 'PumpPressure', 'BlockHeight', 'PumpSpm', 'PumpSpm2', 'PumpSpm3', 'RotaryTorque', 'TopDrive RPM',
              'TopDrive Torque', 'WOB',  'ROP-F','T-HL','BitPosition', 'BitStatus', 'SlipStatus',  'Comments', 'Next 24 Hr Comments', 'Report Id', 'Report Date']

    df = pd.DataFrame(tmpAllJobs)
    df.to_excel(writer, sheet_name='Jobs Processed', index=False)

    df = pd.DataFrame(RAEJobs)
    df = df.sort_values(by=df.columns[0])
    df.to_excel(writer, sheet_name='RAE Report', index=False, header = header)  # header = header

    # df = pd.DataFrame(MR_Report_Ids)
    # df.to_excel(writer, sheet_name='Report_Ids', index=False, header=['Job Ids', 'Report IDs'])

    header2 = ['JobID', 'Report Id', 'Report Date', 'Activity Code', 'Details of Operation']

    df = pd.DataFrame(MR_Report_Comments)
    df = df.sort_values(by=df.columns[0])
    df.to_excel(writer, sheet_name='Report_Comment', index=False, header = header2)  # , header = header2

    df = pd.DataFrame(processedJobList)
    df.to_excel(writer, sheet_name='Jobs not found', index=False, )

    df = pd.DataFrame(ZeroReportList)
    df.to_excel(writer, sheet_name='No Report List', index=False, )

    df = pd.DataFrame(jobsTimeBased)
    df.to_excel(writer, sheet_name='Time Based Pull', index=False, )


    # List of sheet names to hide
    sheets_to_hide = ['Jobs Processed',  'Jobs not found', 'No Report List'] #'Report_Ids',

    # Loop through the list of sheets to hide
    for sheet_name in sheets_to_hide:
        # Get the sheet object
        sheet_to_hide = writer.sheets[sheet_name]
        # Hide the sheet
        sheet_to_hide.sheet_state = 'hidden'
    writer.close()

    wb = openpyxl.load_workbook(f'RAEJobsListTest {today} .xlsx')
    for worksheet in wb.worksheets:
        print(worksheet.title)

    # Define the conditional formatting

    # Select the "RAE Jobs" worksheet
    ws = wb['RAE Report']

    from openpyxl import Workbook
    from openpyxl.formatting.rule import IconSetRule
    from openpyxl.styles import Font, Color

    icon_rule = IconSetRule('3Symbols', 'num', [0, 2, 3])
    icon_rule.dxfId = 0

    # Set font color to white (same as cell background color)
    white_font = Font(color=Color(rgb="FFFFFFFF"))

    for row in ws['H2:T150']:
        for cell in row:
            cell.font = white_font



    ws.conditional_formatting.add('H2:T150', icon_rule)
    wb.save(f'RAEJobsListTest {today} .xlsx')
    wb.close()

    # Sending Email
    from EmailModule import send_email
    recipients = CFG['emailRecipients']
    for recipient in recipients:
        email_body = f'Hi {str(recipient)},\n \nPlease see the attached email for today"s RAE Data\n\n\nAlso, this is a test email.\n There"s no report Date, so I added the report completion status.'
        # send email
        send_email('Test Email From Elvis ', email_body, recipient, f'RAEJobsListTest {today} .xlsx')

    logging.info('Starting script')
    quit()

schedule.every().day.at("06:20").do(main)
if __name__ == "__main__":
    main()

while True:
    schedule.run_pending()
    time.sleep(30)

